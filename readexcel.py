import openpyxl
workbook = openpyxl.load_workbook("C:\\Users\\dell\\Downloads\\testexcel.xlsx")
sheetName =workbook.active
#
# cell = sheetName.cell(2,1)
# print(cell.value)
#
# print(sheetName.max_row)
# print(sheetName.max_column)

# for m in range(1,sheetName.max_row+1):
#     for n in range(1,sheetName.max_column+1):
#         print(sheetName.cell(m,n).value)


# workbook.save("C:\\Users\\dell\\Downloads\\excel.xlsx")

for r in range(1,sheetName.max_row+1):
    name = sheetName.cell(r,1).value
    if name == "kalai1":
        print(sheetName.cell(r,2).value)
